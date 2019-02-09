from openpyxl import Workbook
from openpyxl import load_workbook



wb = Workbook()

wb = load_workbook(filename = 'rawData.xlsx')
ws = wb.active

for i in range(1, 60):
    for j in range(1, 60):
        d =ws.cell(row = i, column = j).value
        if type(d) is str:
            if "plusORMinus" in d:
                ind = d.find("p")
                ws.cell(row=i, column=j).value = d[:ind-1]

wb.save(filename = "fixed.xlsx")

